import sys
import os
from openpyxl import load_workbook, Workbook

def get_longest_common_prefix(strings):
    if not strings:
        return ''
    # Convert all strings to unicode if needed
    prefix = strings[0]
    for s in strings[1:]:
        # Find the common prefix between prefix and s
        i = 0
        while i < min(len(prefix), len(s)) and prefix[i] == s[i]:
            i += 1
        prefix = prefix[:i]
        if not prefix:
            break
    return prefix

def sanitize_sheet_name(name):
    # Remove invalid characters from sheet name
    invalid_chars = '[]:*?/\\'
    for char in invalid_chars:
        name = name.replace(char, '')
    return name

def main():
    # 获取输入文件列表
    if len(sys.argv) < 2:
        print("用法: python script.py <输入文件1.xlsx> <输入文件2.xlsx> ...")
        sys.exit(1)

    input_files = sys.argv[1:]

    # 检查输入文件是否存在
    for input_file in input_files:
        if not os.path.isfile(input_file):
            print(f"文件未找到: {input_file}")
            sys.exit(1)

    # 获取所有文件的文件名
    filenames = [os.path.basename(f) for f in input_files]
    filename_prefix = get_longest_common_prefix(filenames)

    # 判断是否有匹配的前缀
    if filename_prefix:
        output_filename = f"{filename_prefix}合并表.xlsx"
    else:
        output_filename = "合并表.xlsx"

    # 输出文件路径与第一个输入文件在同一目录
    output_dir = os.path.dirname(input_files[0])
    output_path = os.path.join(output_dir, output_filename)

    # 创建新的工作簿
    output_wb = Workbook()
    # 删除默认创建的第一个工作表
    output_wb.remove(output_wb.active)

    # 处理每个输入文件
    for input_file in input_files:
        try:
            wb = load_workbook(input_file)
        except Exception as e:
            print(f"加载工作簿时出错: {input_file}, 错误: {e}")
            continue

        # 获取文件名（不含扩展名）
        file_basename = os.path.splitext(os.path.basename(input_file))[0]

        # 移除前缀以获取剩余文件名
        remaining_name = file_basename.replace(filename_prefix, '', 1).strip() if filename_prefix else file_basename

        # 获取并按升序排序所有子表名
        sorted_sheet_names = sorted(wb.sheetnames)

        for sheet_name in sorted_sheet_names:
            ws = wb[sheet_name]

            # 生成新的子表名
            if filename_prefix:
                new_sheet_name = f"{remaining_name}_{sheet_name}"
            else:
                new_sheet_name = f"{file_basename}_{sheet_name}"

            new_sheet_name = sanitize_sheet_name(new_sheet_name)

            # 如果工作表名已存在，添加数字后缀
            original_sheet_name = new_sheet_name
            counter = 1
            while new_sheet_name in output_wb.sheetnames:
                new_sheet_name = f"{original_sheet_name}_{counter}"
                counter += 1

            # 创建新的工作表
            new_ws = output_wb.create_sheet(title=new_sheet_name)

            # 复制数据
            for row in ws.iter_rows(values_only=False):
                new_row = []
                for cell in row:
                    new_row.append(cell.value)
                new_ws.append(new_row)

    # 保存合并后的工作簿
    try:
        output_wb.save(output_path)
        print(f"已保存合并表: {output_path}")
    except Exception as e:
        print(f"保存工作簿时出错: {output_path}, 错误: {e}")

if __name__ == "__main__":
    main()
