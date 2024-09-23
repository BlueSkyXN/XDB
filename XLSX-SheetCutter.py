import sys
import os
from openpyxl import load_workbook, Workbook

def sanitize_filename(filename):
    # 去除文件名中的非法字符
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        filename = filename.replace(char, '')
    # 去除首尾空格
    filename = filename.strip()
    return filename

def main():
    # 检查是否提供了输入文件路径
    if len(sys.argv) < 2:
        print("用法: python script.py <输入文件.xlsx>")
        sys.exit(1)

    input_file = sys.argv[1]

    # 检查输入文件是否存在
    if not os.path.isfile(input_file):
        print(f"文件未找到: {input_file}")
        sys.exit(1)

    # 获取输入文件的目录和名称
    input_dir = os.path.dirname(input_file)
    input_name = os.path.splitext(os.path.basename(input_file))[0]

    # 加载工作簿
    try:
        wb = load_workbook(input_file)
    except Exception as e:
        print(f"加载工作簿时出错: {e}")
        sys.exit(1)

    # 遍历所有子表
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 创建新的工作簿
        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = sheet_name

        # 复制数据到新的工作簿
        for row in ws.iter_rows(values_only=False):
            new_row = [cell.value for cell in row]
            new_ws.append(new_row)

        # 生成输出文件名，注意处理非法字符
        sanitized_sheet_name = sanitize_filename(sheet_name)
        if not sanitized_sheet_name:
            sanitized_sheet_name = "Sheet"
        output_file = os.path.join(input_dir, f"{input_name}_{sanitized_sheet_name}.xlsx")

        # 保存新的工作簿
        try:
            new_wb.save(output_file)
            print(f"已保存 {output_file}")
        except Exception as e:
            print(f"保存工作簿时出错 {output_file}: {e}")

if __name__ == "__main__":
    main()
