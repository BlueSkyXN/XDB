#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
XDB - 通用Excel/CSV转SQLite/MySQL工具

功能：将任意格式的Excel文件（XLSX）或CSV文件转换为SQLite或MySQL数据库，支持多工作表处理。
特点：高性能并行处理、智能类型映射、自动索引创建、多数据库支持、灵活模式选择。

支持输入格式：
- Excel (XLSX): 支持多工作表、合并单元格等复杂格式
- CSV: 支持不同分隔符、编码和引号规则

支持数据库：
- SQLite: 本地文件数据库
- MySQL: 远程/本地MySQL服务器

支持模式：
- 覆盖模式：删除并重建目标表
- 追加模式：保留表结构，追加数据

字段处理：
- 创建所有：自动创建所有字段
- 匹配目标：仅处理目标数据库已有的字段
"""

import os
import sys
import time
import logging
import argparse
import sqlite3
import pandas as pd
import concurrent.futures
import multiprocessing
import gc
from tqdm import tqdm
import re
import psutil
from datetime import datetime
from openpyxl import load_workbook
import getpass
import pymysql
from contextlib import contextmanager
from abc import ABC, abstractmethod
import csv
import chardet

# 配置日志
def setup_logger(level=logging.INFO, log_file=None):
    """配置日志系统"""
    logger = logging.getLogger()
    logger.setLevel(level)
    
    # 清除已有处理器
    for handler in logger.handlers[:]:
        logger.removeHandler(handler)
    
    # 创建格式化器
    formatter = logging.Formatter(
        '%(asctime)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    
    # 添加控制台处理器
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)
    
    # 添加文件处理器(如果指定)
    if log_file:
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)
    
    return logger

# 检测文件类型
def detect_file_type(file_path):
    """检测文件类型（Excel或CSV）"""
    logger = logging.getLogger(__name__)
    
    _, ext = os.path.splitext(file_path)
    ext = ext.lower()
    
    if ext in ['.xlsx', '.xls']:
        return 'excel'
    elif ext == '.csv':
        return 'csv'
    else:
        # 尝试通过文件内容识别CSV
        try:
            with open(file_path, 'rb') as f:
                sample = f.read(4096)  # 读取前4KB
                
            # 检测是否有逗号或分号等常见CSV分隔符
            text_sample = sample.decode('utf-8', errors='ignore')
            comma_count = text_sample.count(',')
            semicolon_count = text_sample.count(';')
            tab_count = text_sample.count('\t')
            
            # 判断是否可能是CSV
            if comma_count > 5 or semicolon_count > 5 or tab_count > 5:
                logger.info(f"文件 {file_path} 看起来像是CSV格式，将作为CSV处理")
                return 'csv'
                
            logger.warning(f"无法识别 {file_path} 的文件类型，将尝试作为Excel处理")
            return 'excel'
        except Exception as e:
            logger.warning(f"检测文件类型时出错: {str(e)}，将尝试作为Excel处理")
            return 'excel'

# 检测CSV文件的编码和分隔符
def detect_csv_properties(csv_path, sample_size=4096):
    """检测CSV文件的编码和分隔符"""
    logger = logging.getLogger(__name__)
    
    try:
        # 读取文件样本用于检测
        with open(csv_path, 'rb') as f:
            sample = f.read(sample_size)
        
        # 检测编码
        result = chardet.detect(sample)
        encoding = result['encoding'] if result['confidence'] > 0.7 else 'utf-8'
        
        # 使用检测到的编码读取文件样本
        text_sample = sample.decode(encoding, errors='ignore')
        
        # 计算可能的分隔符
        sep_candidates = {
            ',': text_sample.count(','),
            ';': text_sample.count(';'),
            '\t': text_sample.count('\t'),
            '|': text_sample.count('|')
        }
        
        # 确定最有可能的分隔符
        if any(sep_candidates.values()):
            max_sep = max(sep_candidates, key=sep_candidates.get)
        else:
            max_sep = ','  # 默认使用逗号分隔符
        
        # 检查是否有引号字符
        double_quote_count = text_sample.count('"')
        single_quote_count = text_sample.count("'")
        quotechar = '"' if double_quote_count > single_quote_count else "'"
        
        logger.info(f"CSV检测结果 - 编码: {encoding}, 分隔符: {repr(max_sep)}, 引号字符: {quotechar}")
        
        # 尝试检测是否有表头
        lines = text_sample.split('\n')
        has_header = True
        if len(lines) >= 2:
            # 检查第一行和第二行的分隔符计数是否一致
            header_sep_count = lines[0].count(max_sep)
            data_sep_count = lines[1].count(max_sep)
            
            if header_sep_count != data_sep_count:
                has_header = False
                logger.warning("CSV文件可能没有表头，将使用默认列名")
        
        return {
            'encoding': encoding,
            'sep': max_sep,
            'quotechar': quotechar,
            'has_header': has_header
        }
    except Exception as e:
        logger.warning(f"检测CSV属性时出错: {str(e)}，将使用默认设置")
        return {
            'encoding': 'utf-8',
            'sep': ',',
            'quotechar': '"',
            'has_header': True
        }

# 数据类型检测
def detect_column_types(sample_data, headers, db_type='sqlite'):
    """分析样本数据，检测每列的数据类型，根据目标数据库类型返回适当的类型定义"""
    logger = logging.getLogger(__name__)
    
    if not sample_data or not headers:
        return ['TEXT'] * len(headers) if db_type == 'sqlite' else ['VARCHAR(255)'] * len(headers)
    
    detected_types = []
    
    # 检查第一列是否可以作为主键（值唯一）
    first_col_values = [
        row[0] for row in sample_data 
        if row and len(row) > 0 and row[0] is not None
    ]
    is_potential_pk = (len(first_col_values) == len(set(first_col_values)) 
                       and len(first_col_values) > 0)
    
    # 转置样本数据，按列分析
    max_cols = max(len(row) for row in sample_data if row)
    columns = [[] for _ in range(max_cols)]
    
    for row in sample_data:
        for i, val in enumerate(row):
            if i < max_cols and val is not None:
                columns[i].append(val)
    
    for i, column_data in enumerate(columns):
        try:
            # 移除空值
            non_empty = [x for x in column_data if x is not None and str(x).strip() != '']
            if not non_empty:
                # 默认字符串类型
                column_type = 'TEXT' if db_type == 'sqlite' else 'VARCHAR(255)'
            # 整数判断
            elif all(isinstance(x, int) or (isinstance(x, str) and x.isdigit()) for x in non_empty):
                if db_type == 'sqlite':
                    column_type = ('INTEGER PRIMARY KEY' if i == 0 and is_potential_pk else 'INTEGER')
                else:  # MySQL
                    column_type = ('INT AUTO_INCREMENT PRIMARY KEY' if i == 0 and is_potential_pk else 'INT')
            # 浮点数判断
            elif all(isinstance(x, float) or (isinstance(x, str) and 
                    re.match(r'^-?\d+(\.\d+)?$', str(x))) for x in non_empty):
                column_type = 'REAL' if db_type == 'sqlite' else 'DOUBLE'
            # 日期判断
            elif all(isinstance(x, datetime) or (isinstance(x, str) and 
                    any(re.match(pattern, str(x)) for pattern in 
                        [r'\d{4}-\d{1,2}-\d{1,2}', r'\d{1,2}/\d{1,2}/\d{4}'])) 
                    for x in non_empty):
                column_type = 'DATE' if db_type == 'sqlite' else 'DATETIME'
            # 布尔值判断
            elif all(str(x).lower() in ('true', 'false', '0', '1', 'yes', 'no') for x in non_empty):
                column_type = 'BOOLEAN' if db_type == 'sqlite' else 'TINYINT(1)'
            # 长文本判断
            elif any(isinstance(x, str) and len(x) > 255 for x in non_empty):
                column_type = 'TEXT' if db_type == 'sqlite' else 'TEXT'
            else:
                # 默认文本类型
                max_length = max(len(str(x)) for x in non_empty) if non_empty else 255
                max_length = min(max(max_length + 50, 100), 255)  # 为字符串预留一些空间，但不超过255
                column_type = 'TEXT' if db_type == 'sqlite' else f'VARCHAR({max_length})'
                
            detected_types.append(column_type)
        except Exception as e:
            logger.warning(f"列 '{headers[i] if i < len(headers) else i}' 类型检测失败: {str(e)}，使用默认文本类型")
            column_type = 'TEXT' if db_type == 'sqlite' else 'VARCHAR(255)'
            detected_types.append(column_type)
    
    # 补齐类型（如果样本数据列数不足）
    while len(detected_types) < len(headers):
        column_type = 'TEXT' if db_type == 'sqlite' else 'VARCHAR(255)'
        detected_types.append(column_type)
    
    return detected_types

# Excel读取与处理
def get_excel_info(excel_path, sheet_name=None):
    """获取Excel文件的基本信息"""
    logger = logging.getLogger(__name__)
    
    try:
        # 获取所有工作表
        xls = pd.ExcelFile(excel_path)
        sheet_names = xls.sheet_names
        
        if sheet_name is not None and sheet_name not in sheet_names:
            if isinstance(sheet_name, int) and 0 <= sheet_name < len(sheet_names):
                sheet_name = sheet_names[sheet_name]
            else:
                logger.warning(f"未找到工作表 '{sheet_name}'，使用第一个工作表")
                sheet_name = sheet_names[0]
        elif sheet_name is None:
            sheet_name = sheet_names[0]
        
        # 使用openpyxl获取工作表信息，避免加载全部数据
        wb = load_workbook(excel_path, read_only=True)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        
        # 获取表头行
        header_row = next(ws.rows)
        header = [cell.value for cell in header_row]
        
        # 清理表头
        cleaned_header = []
        for i, col in enumerate(header):
            if col is None or str(col).strip() == '':
                col = f"Column_{i+1}"
            else:
                # 替换非法字符
                col = str(col).strip().replace('\n', '_').replace('\r', '_')
                col = ''.join(c if c.isalnum() or c in ['_', '.'] else '_' for c in col)
            
            # 确保唯一性
            while col in cleaned_header:
                col = f"{col}_dup"
            cleaned_header.append(col)
        
        # 估计行数
        try:
            estimated_rows = ws.max_row - 1
        except:
            # 如果max_row不可靠，使用文件大小估算
            file_size_mb = os.path.getsize(excel_path) / (1024*1024)
            estimated_rows = int(file_size_mb * 2000)  # 粗略估算每MB约2000行
        
        wb.close()
        
        return {
            'sheet_names': sheet_names,
            'current_sheet': sheet_name,
            'headers': cleaned_header,
            'raw_headers': header,
            'estimated_rows': estimated_rows
        }
    except Exception as e:
        logger.error(f"获取Excel信息出错: {str(e)}")
        raise

def get_csv_info(csv_path, csv_props=None):
    """获取CSV文件的基本信息"""
    logger = logging.getLogger(__name__)
    
    try:
        # 如果未提供CSV属性，则进行检测
        if csv_props is None:
            csv_props = detect_csv_properties(csv_path)
        
        encoding = csv_props['encoding']
        sep = csv_props['sep']
        quotechar = csv_props['quotechar']
        has_header = csv_props['has_header']
        
        # 获取表头和行数估计
        with open(csv_path, 'r', encoding=encoding, errors='replace') as f:
            # 读取前几行用于表头检测
            sample_lines = []
            for _ in range(5):  # 读取前5行
                line = f.readline()
                if not line:
                    break
                sample_lines.append(line)
            
            # 使用csv模块解析第一行获取字段数
            sniffer = csv.Sniffer()
            dialect = csv.excel
            try:
                # 尝试检测CSV方言
                if sample_lines:
                    dialect = sniffer.sniff('\n'.join(sample_lines), delimiters=',;\t|')
            except (csv.Error, ValueError):
                # 使用检测到的分隔符创建自定义方言
                class CustomDialect(csv.Dialect):
                    delimiter = sep
                    quotechar = quotechar
                    doublequote = True
                    skipinitialspace = True
                    lineterminator = '\r\n'
                    quoting = csv.QUOTE_MINIMAL
                
                dialect = CustomDialect
            
            # 重置文件指针
            f.seek(0)
            
            # 使用csv模块读取表头
            reader = csv.reader(f, dialect)
            if has_header:
                try:
                    raw_headers = next(reader)
                except StopIteration:
                    # 如果文件为空，提供一个默认表头
                    raw_headers = ["Column_1"]
            else:
                # 读取第一行数据以获取列数
                try:
                    first_row = next(reader)
                    # 生成默认列名
                    raw_headers = [f"Column_{i+1}" for i in range(len(first_row))]
                except StopIteration:
                    # 如果文件为空，提供一个默认表头
                    raw_headers = ["Column_1"]
            
            # 清理表头
            cleaned_header = []
            for i, col in enumerate(raw_headers):
                if col is None or str(col).strip() == '':
                    col = f"Column_{i+1}"
                else:
                    # 替换非法字符
                    col = str(col).strip().replace('\n', '_').replace('\r', '_')
                    col = ''.join(c if c.isalnum() or c in ['_', '.'] else '_' for c in col)
                
                # 确保唯一性
                while col in cleaned_header:
                    col = f"{col}_dup"
                cleaned_header.append(col)
            
            # 估计行数
            # 先重置文件指针
            f.seek(0)
            
            # 估算总行数
            file_size = os.path.getsize(csv_path)
            if file_size > 5*1024*1024:  # 如果文件大于5MB，使用估算方法
                # 计算前1000行的平均行长度
                lines_to_check = 1000
                total_chars = 0
                line_count = 0
                for i, line in enumerate(f):
                    if not line:
                        continue
                    total_chars += len(line)
                    line_count += 1
                    if i >= lines_to_check-1:
                        break
                
                if line_count > 0:
                    avg_line_length = total_chars / line_count
                    estimated_rows = int(file_size / avg_line_length)
                else:
                    estimated_rows = 0
            else:
                # 小文件直接计数
                with open(csv_path, 'r', encoding=encoding, errors='replace') as count_f:
                    line_count = sum(1 for _ in count_f)
                    estimated_rows = line_count - (1 if has_header else 0)
        
        # 包装返回结果，模拟Excel信息格式
        return {
            'sheet_names': ['Sheet1'],  # CSV视为单一工作表
            'current_sheet': 'Sheet1',
            'headers': cleaned_header,
            'raw_headers': raw_headers,
            'estimated_rows': estimated_rows,
            'csv_props': csv_props  # 额外返回CSV特有属性
        }
    except Exception as e:
        logger.error(f"获取CSV信息出错: {str(e)}")
        raise

def get_sample_data(file_path, sheet_name=None, sample_size=100, file_type=None, csv_props=None):
    """获取样本数据用于类型检测，支持Excel和CSV"""
    logger = logging.getLogger(__name__)
    
    try:
        # 检测文件类型（如果未提供）
        if file_type is None:
            file_type = detect_file_type(file_path)
        
        if file_type == 'excel':
            # Excel处理逻辑
            df_sample = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                nrows=sample_size + 1  # 读取部分数据
            )
        else:  # CSV文件
            # 如果未提供CSV属性，则进行检测
            if csv_props is None:
                csv_props = detect_csv_properties(file_path)
            
            # 使用pandas读取CSV
            df_sample = pd.read_csv(
                file_path,
                sep=csv_props['sep'],
                encoding=csv_props['encoding'],
                quotechar=csv_props['quotechar'],
                nrows=sample_size + 1,  # 读取部分数据
                header=0 if csv_props['has_header'] else None,
                engine='python'  # 使用python引擎以处理复杂CSV
            )
            
            # 如果没有表头，添加默认列名
            if not csv_props['has_header']:
                df_sample.columns = [f"Column_{i+1}" for i in range(len(df_sample.columns))]
        
        # 移除完全为空的行
        df_sample = df_sample.dropna(how='all')
        
        # 直接返回所有数据行（DataFrame里已经不包含表头行那一行）
        return df_sample.values.tolist(), df_sample.columns.tolist()
    except Exception as e:
        logger.error(f"获取样本数据失败: {str(e)}")
        raise

def get_merged_cells_info(file_path, sheet_name, file_type=None):
    """获取合并单元格信息，仅适用于Excel文件"""
    logger = logging.getLogger(__name__)
    
    try:
        # 检测文件类型（如果未提供）
        if file_type is None:
            file_type = detect_file_type(file_path)
        
        if file_type != 'excel':
            # CSV文件没有合并单元格
            return []
        
        # 读取工作簿以获取合并单元格信息
        wb = load_workbook(file_path, read_only=False, data_only=True)
        ws = wb[sheet_name]
        
        # 获取所有合并单元格范围
        merged_ranges = []
        for merged_cell_range in ws.merged_cells.ranges:
            min_row, min_col, max_row, max_col = (
                merged_cell_range.min_row, 
                merged_cell_range.min_col, 
                merged_cell_range.max_row,
                merged_cell_range.max_col
            )
            merged_ranges.append((min_row, min_col, max_row, max_col))
        
        wb.close()
        
        return merged_ranges
    except Exception as e:
        logger.warning(f"获取合并单元格信息失败: {str(e)}")
        return []

def process_chunk(args):
    """处理单个数据块的函数(用于并行处理)，支持Excel和CSV"""
    chunk_id, file_path, sheet_name, skiprows, nrows, headers, merged_ranges, file_type, csv_props, transform_rules = args
    # 注意:这里增加了transform_rules参数!
    
    # 每个进程使用独立的日志记录器
    worker_logger = logging.getLogger(f"worker_{os.getpid()}")
    handler = logging.StreamHandler()
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    handler.setFormatter(formatter)
    worker_logger.addHandler(handler)
    worker_logger.setLevel(logging.INFO)
    
    try:
        if file_type == 'excel':
            # Excel文件处理
            df_chunk = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                skiprows=skiprows,
                nrows=nrows,
                header=None
            )
            
            # 处理合并单元格 - 复制值到所有被合并的单元格位置
            if merged_ranges:
                for min_row, min_col, max_row, max_col in merged_ranges:
                    # 检查合并区域是否在当前块内
                    chunk_min_row = min_row - skiprows
                    chunk_max_row = max_row - skiprows
                    
                    # 如果合并区域与当前数据块有交集
                    if chunk_min_row < df_chunk.shape[0] and chunk_max_row >= 0:
                        # 调整相对于当前块的行索引
                        chunk_min_row = max(0, chunk_min_row)
                        chunk_max_row = min(df_chunk.shape[0] - 1, chunk_max_row)
                        
                        # 调整列索引(pandas是0-based，而openpyxl是1-based)
                        df_min_col = min_col - 1
                        df_max_col = min(df_chunk.shape[1] - 1, max_col - 1)
                        
                        if chunk_min_row >= 0 and df_min_col >= 0:
                            try:
                                cell_value = df_chunk.iloc[chunk_min_row, df_min_col]
                                
                                for r in range(chunk_min_row, chunk_max_row + 1):
                                    for c in range(df_min_col, df_max_col + 1):
                                        if r < df_chunk.shape[0] and c < df_chunk.shape[1]:
                                            df_chunk.iloc[r, c] = cell_value
                            except:
                                pass
        else:  # CSV文件处理
            # 使用pandas读取CSV块
            df_chunk = pd.read_csv(
                file_path,
                sep=csv_props['sep'],
                encoding=csv_props['encoding'],
                quotechar=csv_props['quotechar'],
                skiprows=skiprows,
                nrows=nrows,
                header=None,
                engine='python'  # 使用python引擎以处理复杂CSV
            )
        
        # 移除完全为空的行
        df_chunk = df_chunk.dropna(how='all')
        
        # 处理数据
        processed_data = []
        for row in df_chunk.values.tolist():
            # 确保行长度与列头匹配
            processed_row = []
            for i in range(len(headers)):
                if i < len(row):
                    cell_value = row[i]
                else:
                    cell_value = None
                
                # 应用字段转换规则
                if transform_rules and headers[i] in transform_rules:
                    cell_value = apply_column_transformation(cell_value, transform_rules[headers[i]])
                
                # 处理 NaN 和空值
                if cell_value is None or (isinstance(cell_value, float) and cell_value != cell_value) or (isinstance(cell_value, str) and cell_value.strip() == ''):
                    processed_row.append(None)
                else:
                    if isinstance(cell_value, str):
                        cell_value = cell_value.replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ')
                    processed_row.append(cell_value)
            
            processed_data.append(tuple(processed_row))
        
        return chunk_id, processed_data
    except Exception as e:
        worker_logger.error(f"处理块 {chunk_id} 出错: {str(e)}")
        return chunk_id, []

# 数据库抽象基类
class Database(ABC):
    """数据库操作抽象基类"""
    
    @abstractmethod
    def connect(self):
        """连接到数据库"""
        pass
    
    @abstractmethod
    def disconnect(self):
        """断开数据库连接"""
        pass
    
    @abstractmethod
    def table_exists(self, table_name):
        """检查表是否存在"""
        pass
    
    @abstractmethod
    def drop_table(self, table_name):
        """删除表"""
        pass
    
    @abstractmethod
    def get_table_columns(self, table_name):
        """获取表的列信息"""
        pass
    
    @abstractmethod
    def create_table(self, table_name, headers, column_types, has_pk=False):
        """创建表"""
        pass
    
    @abstractmethod
    def write_data(self, table_name, headers, data_chunks, field_mapping=None):
        """写入数据"""
        pass
    
    @abstractmethod
    def create_indices(self, table_name, headers, max_indices=3):
        """创建索引"""
        pass
    
    @abstractmethod
    def optimize(self):
        """优化数据库"""
        pass
    
    @abstractmethod
    def verify(self):
        """验证数据库"""
        pass

# SQLite数据库实现
class SQLiteDatabase(Database):
    """SQLite数据库操作实现"""
    
    def __init__(self, db_path):
        """初始化"""
        self.db_path = db_path
        self.conn = None
        self.logger = logging.getLogger(__name__)
    
    def connect(self):
        """连接到SQLite数据库"""
        try:
            # 确保输出目录存在（如果真的有目录）
            db_dir = os.path.dirname(self.db_path)
            if db_dir:
                os.makedirs(db_dir, exist_ok=True)
            
            # 创建连接
            self.conn = sqlite3.connect(self.db_path)
            
            # 优化设置
            self.conn.execute('PRAGMA journal_mode = WAL')
            self.conn.execute('PRAGMA synchronous = NORMAL')
            self.conn.execute('PRAGMA cache_size = 100000')
            self.conn.execute('PRAGMA temp_store = MEMORY')
            
            # 动态设置内存映射大小
            available_memory = psutil.virtual_memory().available
            mmap_size = min(int(available_memory * 0.5), 30000000000)  
            self.conn.execute(f'PRAGMA mmap_size = {mmap_size}')
            
            self.logger.info(f"已连接到SQLite数据库: {self.db_path}")
            self.logger.info(f"已设置内存映射大小: {mmap_size / (1024**3):.1f} GB")
            
            return self.conn
        except Exception as e:
            self.logger.error(f"连接SQLite数据库出错: {str(e)}")
            raise
    
    def disconnect(self):
        """断开SQLite连接"""
        if self.conn:
            self.conn.close()
            self.conn = None
            self.logger.info("已断开SQLite数据库连接")
    
    def table_exists(self, table_name):
        """检查表是否存在"""
        if not self.conn:
            self.connect()
        
        cursor = self.conn.cursor()
        cursor.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
        return cursor.fetchone() is not None
    
    def drop_table(self, table_name):
        """删除表"""
        if not self.conn:
            self.connect()
        
        try:
            self.conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')
            self.conn.commit()
            self.logger.info(f"已删除表: {table_name}")
            return True
        except Exception as e:
            self.logger.error(f"删除表 {table_name} 出错: {str(e)}")
            return False
    
    def get_table_columns(self, table_name):
        """获取表的列信息"""
        if not self.conn:
            self.connect()
        
        try:
            cursor = self.conn.cursor()
            cursor.execute(f"PRAGMA table_info(\"{table_name}\")")
            columns = cursor.fetchall()
            
            # SQLite PRAGMA table_info 返回的格式: (cid, name, type, notnull, dflt_value, pk)
            column_info = {col[1]: {'type': col[2], 'position': col[0]} for col in columns}
            return column_info
        except Exception as e:
            self.logger.error(f"获取表 {table_name} 的列信息出错: {str(e)}")
            return {}
    
    # 修改SQLiteDatabase.create_table方法
    def create_table(self, table_name, headers, column_types, has_pk=False):
        """创建SQLite表"""
        if not self.conn:
            self.connect()
        
        try:
            column_defs = []
            
            if has_pk:
                # 第一列已是主键
                column_defs.append(f'"{headers[0]}" {column_types[0]}')
                
                # 添加其他列（从索引1开始，类型也从索引1开始）
                for i in range(1, len(headers)):
                    if i < len(column_types):
                        safe_column = f'"{headers[i]}"'
                        column_defs.append(f'{safe_column} {column_types[i]}')
            else:
                # 添加自增ID主键
                column_defs.append('id INTEGER PRIMARY KEY AUTOINCREMENT')
                
                # 添加所有列（从索引0开始）
                for i in range(len(headers)):
                    if i < len(column_types):
                        safe_column = f'"{headers[i]}"'
                        column_defs.append(f'{safe_column} {column_types[i]}')
            
            # 创建表
            create_table_sql = f'''
            CREATE TABLE IF NOT EXISTS "{table_name}" (
                {', '.join(column_defs)}
            )
            '''
            
            self.conn.execute(create_table_sql)
            self.conn.commit()
            
            self.logger.info(f"已创建SQLite表 '{table_name}' 包含 {len(headers)} 列")
            return True
        except Exception as e:
            self.logger.error(f"创建SQLite表出错: {str(e)}")
            raise

    def write_data(self, table_name, headers, data_chunks, field_mapping=None):
        """将处理好的数据块写入SQLite数据库"""
        if not self.conn:
            self.connect()
        
        try:
            total_rows = 0
            
            # 如果有字段映射，使用映射后的字段列表
            if field_mapping:
                # 过滤掉不在映射中的字段
                excel_headers = [h for h in headers if h in field_mapping]
                # 获取映射后的数据库列名
                db_headers = [field_mapping[h] for h in excel_headers]
                self.logger.debug(f"字段映射: Excel列名 -> 数据库列名: {dict(zip(excel_headers, db_headers))}")
            else:
                excel_headers = headers
                db_headers = headers
            
            # 准备批量插入的SQL语句 - 使用映射后的数据库列名
            placeholders = ', '.join(['?'] * len(db_headers))
            insert_sql = f'''
            INSERT INTO "{table_name}" (
                {', '.join([f'"{col}"' for col in db_headers])}
            ) VALUES ({placeholders})
            '''
            
            # 开始一个事务
            with self.conn:
                for chunk_data in data_chunks:
                    if chunk_data:
                        # 如果有字段映射，需要重新组织数据
                        if field_mapping:
                            filtered_data = []
                            for row in chunk_data:
                                # 只保留映射中的字段数据，并按映射后的顺序组织
                                filtered_row = []
                                for i, h in enumerate(headers):
                                    if h in field_mapping:
                                        val = row[i]
                                        # 检查是否为 NaN 值
                                        if isinstance(val, float) and (val != val):
                                            filtered_row.append(None)
                                        else:
                                            filtered_row.append(val)
                                filtered_data.append(tuple(filtered_row))
                            self.conn.executemany(insert_sql, filtered_data)
                            total_rows += len(filtered_data)
                        else:
                            self.conn.executemany(insert_sql, chunk_data)
                            total_rows += len(chunk_data)
            
            self.logger.info(f"成功插入 {total_rows} 行数据到SQLite表 '{table_name}'")
            return total_rows
        except Exception as e:
            self.logger.error(f"写入SQLite出错: {str(e)}")
            raise
    
    def create_indices(self, table_name, headers, max_indices=3):
        """在SQLite表上创建索引以提高查询性能"""
        if not self.conn:
            self.connect()
        
        try:
            self.logger.info("创建SQLite索引...")
            
            # 索引关键词
            index_keywords = ['id', 'code', 'name', 'type', 'date', 'time', 'key', 'num', 'uuid', 'no']
            
            # 查找可能适合索引的列
            index_candidates = []
            for i, col in enumerate(headers):
                col_lower = col.lower()
                if any(keyword in col_lower for keyword in index_keywords):
                    index_candidates.append((i, col))
            
            # 限制索引数量
            index_candidates = index_candidates[:max_indices]
            
            # 创建索引
            for idx, (_, col) in enumerate(index_candidates):
                idx_name = f"idx_{table_name}_{idx}"
                self.conn.execute(f'CREATE INDEX IF NOT EXISTS "{idx_name}" ON "{table_name}" ("{col}")')
                self.logger.info(f"创建SQLite索引: {idx_name} 在列 '{col}'")
            
            self.conn.commit()
            self.logger.info(f"为SQLite表 '{table_name}' 创建了 {len(index_candidates)} 个索引")
            return True
        except Exception as e:
            self.logger.error(f"创建SQLite索引出错: {str(e)}")
            return False
    
    def optimize(self):
        """优化SQLite数据库"""
        if not self.conn:
            self.connect()
        
        try:
            self.logger.info("优化SQLite数据库...")
            self.conn.execute('VACUUM')
            self.conn.commit()
            self.logger.info("SQLite数据库优化完成")
            return True
        except Exception as e:
            self.logger.error(f"优化SQLite数据库出错: {str(e)}")
            return False
    
    def verify(self):
        """验证生成的SQLite数据库"""
        if not self.conn:
            self.connect()
        
        try:
            self.logger.info("验证SQLite数据库...")
            cursor = self.conn.cursor()
            
            # 获取所有表
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = cursor.fetchall()
            
            self.logger.info(f"SQLite数据库包含 {len(tables)} 个表:")
            table_stats = []
            
            for table in tables:
                table_name = table[0]
                # 获取行数
                cursor.execute(f"SELECT COUNT(*) FROM \"{table_name}\"")
                row_count = cursor.fetchone()[0]
                
                # 获取列信息
                cursor.execute(f"PRAGMA table_info(\"{table_name}\")")
                columns = cursor.fetchall()
                
                table_info = {
                    'table_name': table_name,
                    'row_count': row_count,
                    'column_count': len(columns)
                }
                table_stats.append(table_info)
                
                self.logger.info(f"表 '{table_name}': {row_count} 行, {len(columns)} 列")
            
            self.logger.info("SQLite数据库验证完成")
            return table_stats
        except Exception as e:
            self.logger.error(f"验证SQLite数据库出错: {str(e)}")
            return []

# MySQL数据库实现
class MySQLDatabase(Database):
    """MySQL数据库操作实现"""
    
    def __init__(self, host, port, user, password, database, charset='utf8mb4'):
        """初始化"""
        self.host = host
        self.port = port
        self.user = user
        self.password = password
        self.database = database
        self.charset = charset
        self.conn = None
        self.logger = logging.getLogger(__name__)
    
    def connect(self):
        """连接到MySQL数据库"""
        try:
            # 创建连接
            self.conn = pymysql.connect(
                host=self.host,
                port=self.port,
                user=self.user,
                password=self.password,
                database=self.database,
                charset=self.charset,
                autocommit=False,
                cursorclass=pymysql.cursors.DictCursor  # 直接在连接参数中指定
            )
            
            # 优化设置 - 适当调整缓冲区大小
            with self.conn.cursor() as cursor:
                cursor.execute("SET autocommit = 0")
                cursor.execute("SET unique_checks = 0")
                cursor.execute("SET foreign_key_checks = 0")
                cursor.execute("SET sql_mode = 'NO_ENGINE_SUBSTITUTION'")
            
            self.logger.info(f"已连接到MySQL数据库: {self.host}:{self.port}/{self.database}")
            return self.conn
        except Exception as e:
            self.logger.error(f"连接MySQL数据库出错: {str(e)}")
            raise

    def disconnect(self):
        """断开MySQL连接"""
        if self.conn:
            # 恢复默认设置
            try:
                with self.conn.cursor() as cursor:
                    cursor.execute("SET autocommit = 1")
                    cursor.execute("SET unique_checks = 1")
                    cursor.execute("SET foreign_key_checks = 1")
                self.conn.commit()
            except:
                pass
                
            self.conn.close()
            self.conn = None
            self.logger.info("已断开MySQL数据库连接")
    
    def table_exists(self, table_name):
        """检查表是否存在"""
        if not self.conn:
            self.connect()
        
        with self.conn.cursor() as cursor:
            cursor.execute(
                "SELECT COUNT(*) as count FROM information_schema.tables "
                "WHERE table_schema = %s AND table_name = %s",
                (self.database, table_name)
            )
            result = cursor.fetchone()
            return result['count'] > 0
    
    def drop_table(self, table_name):
        """删除表"""
        if not self.conn:
            self.connect()
        
        try:
            with self.conn.cursor() as cursor:
                cursor.execute(f"DROP TABLE IF EXISTS `{table_name}`")
            self.conn.commit()
            self.logger.info(f"已删除MySQL表: {table_name}")
            return True
        except Exception as e:
            self.logger.error(f"删除MySQL表 {table_name} 出错: {str(e)}")
            self.conn.rollback()
            return False
    
    def get_table_columns(self, table_name):
        """获取表的列信息"""
        if not self.conn:
            self.connect()
        
        try:
            with self.conn.cursor() as cursor:
                cursor.execute(
                    "SELECT COLUMN_NAME, DATA_TYPE, ORDINAL_POSITION "
                    "FROM INFORMATION_SCHEMA.COLUMNS "
                    "WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s",
                    (self.database, table_name)
                )
                columns = cursor.fetchall()
            
            column_info = {
                col['COLUMN_NAME']: {
                    'type': col['DATA_TYPE'],
                    'position': col['ORDINAL_POSITION'] - 1  # 转为0-based索引
                }
                for col in columns
            }
            return column_info
        except Exception as e:
            self.logger.error(f"获取MySQL表 {table_name} 的列信息出错: {str(e)}")
            return {}
    
    # 修改MySQLDatabase.create_table方法
    def create_table(self, table_name, headers, column_types, has_pk=False):
        """创建MySQL表"""
        if not self.conn:
            self.connect()
        
        try:
            column_defs = []
            
            if has_pk:
                # 第一列已是主键
                column_defs.append(f'`{headers[0]}` {column_types[0]}')
                
                # 添加其他列（从索引1开始，类型也从索引1开始）
                for i in range(1, len(headers)):
                    if i < len(column_types):
                        safe_column = f'`{headers[i]}`'
                        column_defs.append(f'{safe_column} {column_types[i]}')
            else:
                # 添加自增ID主键
                column_defs.append('`id` INT AUTO_INCREMENT PRIMARY KEY')
                
                # 添加所有列（从索引0开始）
                for i in range(len(headers)):
                    if i < len(column_types):
                        safe_column = f'`{headers[i]}`'
                        column_defs.append(f'{safe_column} {column_types[i]}')
            
            # 创建表
            create_table_sql = f'''
            CREATE TABLE IF NOT EXISTS `{table_name}` (
                {', '.join(column_defs)}
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
            '''
            
            with self.conn.cursor() as cursor:
                cursor.execute(create_table_sql)
            self.conn.commit()
            
            self.logger.info(f"已创建MySQL表 '{table_name}' 包含 {len(headers)} 列")
            return True
        except Exception as e:
            self.logger.error(f"创建MySQL表出错: {str(e)}")
            self.conn.rollback()
            raise

    def write_data(self, table_name, headers, data_chunks, field_mapping=None):
        """将处理好的数据块写入MySQL数据库"""
        if not self.conn:
            self.connect()
        
        try:
            total_rows = 0
            
            # 如果有字段映射，使用映射后的字段列表
            if field_mapping:
                # 过滤掉不在映射中的字段
                excel_headers = [h for h in headers if h in field_mapping]
                # 获取映射后的数据库列名
                db_headers = [field_mapping[h] for h in excel_headers]
                self.logger.debug(f"字段映射: Excel列名 -> 数据库列名: {dict(zip(excel_headers, db_headers))}")
            else:
                excel_headers = headers
                db_headers = headers
            
            # 准备批量插入的SQL语句 - 使用映射后的数据库列名
            placeholders = ', '.join(['%s'] * len(db_headers))
            insert_sql = f'''
            INSERT INTO `{table_name}` (
                {', '.join([f'`{col}`' for col in db_headers])}
            ) VALUES ({placeholders})
            '''
            
            # 分批处理，避免过大的事务
            batch_size = 5000  # 每个批次的最大行数
            
            with self.conn.cursor() as cursor:
                for chunk_data in data_chunks:
                    if not chunk_data:
                        continue
                        
                    # 如果有字段映射，需要重新组织数据
                    if field_mapping:
                        filtered_data = []
                        for row in chunk_data:
                            # 只保留映射中的字段数据，并按映射后的顺序组织
                            filtered_row = []
                            for i, h in enumerate(headers):
                                if h in field_mapping:
                                    val = row[i]
                                    # 检查是否为 NaN 值
                                    if isinstance(val, float) and (val != val):
                                        filtered_row.append(None)
                                    else:
                                        filtered_row.append(val)
                            filtered_data.append(tuple(filtered_row))
                    else:
                        # 处理所有字段的 NaN 值
                        filtered_data = []
                        for row in chunk_data:
                            processed_row = []
                            for val in row:
                                if isinstance(val, float) and (val != val):
                                    processed_row.append(None)
                                else:
                                    processed_row.append(val)
                            filtered_data.append(tuple(processed_row))
                    
                    # 分批执行插入
                    for i in range(0, len(filtered_data), batch_size):
                        batch = filtered_data[i:i+batch_size]
                        cursor.executemany(insert_sql, batch)
                        total_rows += len(batch)
                        
                        # 每批次提交一次，减轻数据库压力
                        self.conn.commit()
            
            self.logger.info(f"成功插入 {total_rows} 行数据到MySQL表 '{table_name}'")
            return total_rows
        except Exception as e:
            self.logger.error(f"写入MySQL出错: {str(e)}")
            self.conn.rollback()
            raise
    
    def create_indices(self, table_name, headers, max_indices=3):
        """在MySQL表上创建索引以提高查询性能"""
        if not self.conn:
            self.connect()
        
        try:
            self.logger.info("创建MySQL索引...")
            
            # 索引关键词
            index_keywords = ['id', 'code', 'name', 'type', 'date', 'time', 'key', 'num', 'uuid', 'no']
            
            # 查找可能适合索引的列
            index_candidates = []
            for i, col in enumerate(headers):
                col_lower = col.lower()
                if any(keyword in col_lower for keyword in index_keywords):
                    index_candidates.append((i, col))
            
            # 限制索引数量
            index_candidates = index_candidates[:max_indices]
            
            # 创建索引
            with self.conn.cursor() as cursor:
                for idx, (_, col) in enumerate(index_candidates):
                    idx_name = f"idx_{table_name}_{idx}"
                    # 检查索引是否已存在
                    cursor.execute(
                        "SELECT COUNT(*) as count FROM information_schema.statistics "
                        "WHERE table_schema = %s AND table_name = %s AND index_name = %s",
                        (self.database, table_name, idx_name)
                    )
                    result = cursor.fetchone()
                    
                    if result['count'] == 0:
                        cursor.execute(f"CREATE INDEX `{idx_name}` ON `{table_name}` (`{col}`)")
                        self.logger.info(f"创建MySQL索引: {idx_name} 在列 '{col}'")
            
            self.conn.commit()
            self.logger.info(f"为MySQL表 '{table_name}' 创建了 {len(index_candidates)} 个索引")
            return True
        except Exception as e:
            self.logger.error(f"创建MySQL索引出错: {str(e)}")
            self.conn.rollback()
            return False
    
    def optimize(self):
        """优化MySQL数据库表"""
        if not self.conn:
            self.connect()
        
        try:
            self.logger.info("优化MySQL数据库表...")
            
            # 获取所有表
            with self.conn.cursor() as cursor:
                cursor.execute(
                    "SELECT table_name FROM information_schema.tables "
                    "WHERE table_schema = %s AND table_type = 'BASE TABLE'",
                    (self.database,)
                )
                tables = cursor.fetchall()
                
                for table in tables:
                    table_name = table['table_name']
                    self.logger.info(f"优化表: {table_name}")
                    cursor.execute(f"OPTIMIZE TABLE `{table_name}`")
            
            self.conn.commit()
            self.logger.info("MySQL数据库表优化完成")
            return True
        except Exception as e:
            self.logger.error(f"优化MySQL数据库出错: {str(e)}")
            return False
    
    def verify(self):
        """验证MySQL数据库"""
        if not self.conn:
            self.connect()
        
        try:
            self.logger.info("验证MySQL数据库...")
            
            # 获取所有表
            with self.conn.cursor() as cursor:
                cursor.execute(
                    "SELECT table_name FROM information_schema.tables "
                    "WHERE table_schema = %s AND table_type = 'BASE TABLE'",
                    (self.database,)
                )
                tables = cursor.fetchall()
            
            self.logger.info(f"MySQL数据库包含 {len(tables)} 个表:")
            table_stats = []
            
            for table in tables:
                table_name = table['table_name']
                
                with self.conn.cursor() as cursor:
                    # 获取行数
                    cursor.execute(f"SELECT COUNT(*) as count FROM `{table_name}`")
                    row_count = cursor.fetchone()['count']
                    
                    # 获取列信息
                    cursor.execute(
                        "SELECT COUNT(*) as count FROM information_schema.columns "
                        "WHERE table_schema = %s AND table_name = %s",
                        (self.database, table_name)
                    )
                    column_count = cursor.fetchone()['count']
                
                table_info = {
                    'table_name': table_name,
                    'row_count': row_count,
                    'column_count': column_count
                }
                table_stats.append(table_info)
                
                self.logger.info(f"表 '{table_name}': {row_count} 行, {column_count} 列")
            
            self.logger.info("MySQL数据库验证完成")
            return table_stats
        except Exception as e:
            self.logger.error(f"验证MySQL数据库出错: {str(e)}")
            return []

# 数据库工厂类
class DatabaseFactory:
    """数据库工厂类，用于创建不同类型的数据库连接"""
    
    @staticmethod
    def create_database(db_type, **kwargs):
        """创建数据库连接"""
        if db_type.lower() == 'sqlite':
            return SQLiteDatabase(kwargs.get('db_path'))
        elif db_type.lower() == 'mysql':
            return MySQLDatabase(
                host=kwargs.get('host', 'localhost'),
                port=kwargs.get('port', 3306),
                user=kwargs.get('user', 'root'),
                password=kwargs.get('password', ''),
                database=kwargs.get('database', '')
            )
        else:
            raise ValueError(f"不支持的数据库类型: {db_type}")

# 添加表映射解析函数
def parse_table_mapping(mapping_str):
    """解析表名映射字符串 'excel_sheet=db_table,excel_sheet2=db_table2'"""
    logger = logging.getLogger(__name__)
    
    try:
        mapping = {}
        if not mapping_str:
            return mapping
            
        # 按逗号分割映射对
        mapping_pairs = mapping_str.split(',')
        
        for pair in mapping_pairs:
            if not pair.strip():
                continue
                
            parts = pair.split('=', 1)
            if len(parts) != 2:
                logger.warning(f"跳过无效的表映射: {pair}")
                continue
                
            excel_sheet = parts[0].strip()
            db_table = parts[1].strip()
            
            if not excel_sheet or not db_table:
                logger.warning(f"跳过空的表映射: {pair}")
                continue
                
            mapping[excel_sheet] = db_table
        
        logger.info(f"已解析表映射配置，共 {len(mapping)} 个映射")
        return mapping
        
    except Exception as e:
        logger.error(f"解析表映射出错: {str(e)}")
        return {}

# 转换器主函数
def file_to_database(
    file_path, 
    db_type,
    db_params,
    sheet_name=None, 
    chunk_size=20000, 
    max_workers=None, 
    max_indices=3,
    mode='overwrite',
    field_mode='create-all',
    field_mapping=None,
    column_transforms=None,
    target_table=None,
    table_mapping=None,
    csv_params=None  # 新增参数，用于CSV特定设置
):
    """将Excel或CSV转换为数据库的主函数"""
    logger = logging.getLogger(__name__)
    
    start_time = time.time()
    
    # 如果未指定worker数量，使用CPU核心数减1
    if max_workers is None:
        max_workers = max(1, multiprocessing.cpu_count() - 1)
    
    # 检测文件类型
    file_type = detect_file_type(file_path)
    logger.info(f"检测到文件类型: {file_type}")
    
    # 验证支持的文件类型
    if file_type not in ['excel', 'csv']:
        logger.error(f"不支持的文件类型: {file_type}")
        return {'success': False, 'error': f"不支持的文件类型: {file_type}"}
    
    logger.info(f"开始处理: {file_path} -> {db_type}")
    logger.info(f"使用 {max_workers} 个工作进程，每块 {chunk_size} 行")
    logger.info(f"操作模式: {mode}, 字段处理模式: {field_mode}")
    
    try:
        # 根据文件类型获取文件信息
        csv_props = None
        if file_type == 'excel':
            file_info = get_excel_info(file_path, sheet_name)
            sheet_names = file_info['sheet_names']
            
            # 如果指定了表名，仅处理该表
            if sheet_name is not None:
                sheets_to_process = [file_info['current_sheet']]
            else:
                sheets_to_process = sheet_names
            
            logger.info(f"Excel文件包含 {len(sheet_names)} 个工作表: {', '.join(sheet_names)}")
            logger.info(f"将处理以下工作表: {', '.join(sheets_to_process)}")
        else:  # CSV文件
            # 如果是CSV文件，检测其属性
            if csv_params is not None:
                csv_props = csv_params
            else:
                csv_props = detect_csv_properties(file_path)
            
            file_info = get_csv_info(file_path, csv_props)
            logger.info(f"CSV文件编码: {csv_props['encoding']}, 分隔符: {repr(csv_props['sep'])}")
            
            # CSV文件只有一个工作表
            sheets_to_process = ['Sheet1']
        
        # 创建数据库连接
        db = DatabaseFactory.create_database(db_type, **db_params)
        db.connect()
        
        total_rows_processed = 0
        results = []
        
        for current_sheet in sheets_to_process:
            logger.info(f"开始处理工作表: {current_sheet}")
            sheet_start_time = time.time()
            
            # 获取表头和估计行数
            if file_type == 'excel':
                sheet_info = get_excel_info(file_path, current_sheet)
                headers = sheet_info['headers']
                estimated_rows = sheet_info['estimated_rows']
            else:  # CSV文件
                headers = file_info['headers']
                estimated_rows = file_info['estimated_rows']
            
            # 确定目标表名: 表映射 > 单一目标表 > 工作表名
            if file_type == 'excel' and table_mapping and current_sheet in table_mapping:
                # 优先使用表映射中指定的表名
                table_name = table_mapping[current_sheet]
                logger.info(f"根据映射，工作表 '{current_sheet}' 将导入到表 '{table_name}'")
            elif target_table and (len(sheets_to_process) == 1 or sheet_name == current_sheet):
                # 当只处理一个工作表或当前工作表是指定的工作表时，使用目标表名
                table_name = target_table
                logger.info(f"将使用指定的目标表名 '{table_name}' 而非工作表名 '{current_sheet}'")
            else:
                # 默认使用工作表名（规范化处理）
                table_name = current_sheet.replace(' ', '_').replace('-', '_').replace('.', '_')
                table_name = ''.join(c if c.isalnum() or c == '_' else '_' for c in table_name)
                logger.info(f"使用规范化的工作表名 '{table_name}' 作为目标表名")

            # 确保表名合法性
            table_name = ''.join(c if c.isalnum() or c == '_' else '_' for c in table_name)
            if not table_name or not table_name[0].isalpha():
                table_name = 'T_' + table_name
            
            logger.info(f"工作表 '{current_sheet}' 信息: {len(headers)} 列, 约 {estimated_rows} 行")
            
            # 检查目标表是否存在
            table_exists = db.table_exists(table_name)
            excel_to_db_mapping = None
            used_headers = headers  # 默认使用所有列
            transform_rules = {}    # 字段转换规则
            
            # 处理字段映射和转换规则
            if field_mapping and current_sheet in field_mapping:
                sheet_mapping = field_mapping[current_sheet]
                if sheet_mapping:
                    logger.info(f"使用自定义映射处理工作表 '{current_sheet}'")
                    field_mode = 'mapping'  # 强制使用映射模式
                    
                    # 转换为 {excel列名: 数据库列名} 格式
                    excel_to_db_mapping = sheet_mapping
                    
                    # 获取使用到的Excel列
                    used_headers = list(excel_to_db_mapping.keys())
                    logger.info(f"映射配置包含 {len(used_headers)} 个字段")
            
            # 应用列转换规则
            if column_transforms and current_sheet in column_transforms:
                transform_rules = column_transforms[current_sheet]
                logger.info(f"应用 {len(transform_rules)} 个字段转换规则")
            
            # 处理表和字段映射
            if table_exists:
                if mode == 'overwrite':
                    # 覆盖模式，删除已有表
                    logger.info(f"覆盖模式: 删除已有表 '{table_name}'")
                    db.drop_table(table_name)
                    table_exists = False
                elif field_mode == 'match-only':
                    # 追加模式 + 匹配字段，获取目标表的字段信息
                    logger.info(f"追加模式 + 匹配字段: 获取目标表 '{table_name}' 的字段信息")
                    table_columns = db.get_table_columns(table_name)
                    
                    # 创建字段映射 {源字段名: 目标字段位置}
                    source_to_target_pos = {}
                    for i, header in enumerate(headers):
                        if header in table_columns:
                            source_to_target_pos[header] = table_columns[header]['position']
                    
                    if source_to_target_pos:
                        logger.info(f"找到 {len(source_to_target_pos)} 个匹配的字段")
                        excel_to_db_mapping = {h: h for h in source_to_target_pos.keys()}
                        used_headers = list(excel_to_db_mapping.keys())
                    else:
                        logger.warning(f"目标表 '{table_name}' 没有匹配的字段，将跳过")
                        continue
                elif field_mode == 'mapping' and excel_to_db_mapping:
                    # 映射模式 + 追加，验证目标列是否存在
                    logger.info(f"映射模式 + 追加: 验证目标表 '{table_name}' 中的映射字段")
                    table_columns = db.get_table_columns(table_name)
                    
                    # 检查目标列是否都存在
                    # 注意: 需要使用映射后的数据库列名进行检查
                    target_columns = set(excel_to_db_mapping.values())
                    existing_columns = set(table_columns.keys())
                    missing_columns = target_columns - existing_columns
                    
                    if missing_columns:
                        logger.warning(f"目标表中缺少映射的字段: {', '.join(missing_columns)}")
                        # 过滤掉映射到不存在列的映射
                        excel_to_db_mapping = {
                            excel_col: db_col for excel_col, db_col in excel_to_db_mapping.items()
                            if db_col in existing_columns
                        }
                        
                    if not excel_to_db_mapping:
                        logger.warning(f"映射后没有可用字段，将跳过")
                        continue
                        
                    used_headers = list(excel_to_db_mapping.keys())
            
            # 获取样本数据并检测类型
            if not table_exists:
                sample_data, _ = get_sample_data(file_path, current_sheet if file_type == 'excel' else None, 
                                              file_type=file_type, csv_props=csv_props)
                
                if field_mode == 'mapping' and excel_to_db_mapping:
                    # 使用映射创建表
                    # 从样本数据中提取出映射的列
                    mapped_headers = list(excel_to_db_mapping.values())
                    
                    # 创建一个新的样本数据集，仅包含映射的列
                    mapped_sample = []
                    for row in sample_data:
                        mapped_row = []
                        for i, header in enumerate(headers):
                            if header in excel_to_db_mapping and i < len(row):
                                mapped_row.append(row[i])
                        if mapped_row:  # 只添加非空行
                            mapped_sample.append(mapped_row)
                    
                    column_types = detect_column_types(mapped_sample, mapped_headers, db_type)
                    
                    # 检查是否第一列为主键
                    has_pk = any('PRIMARY KEY' in ct for ct in column_types[:1]) if column_types else False
                    
                    # 创建表结构 - 使用映射后的列名
                    db.create_table(table_name, mapped_headers, column_types, has_pk)
                else:
                    # 常规创建表
                    column_types = detect_column_types(sample_data, headers, db_type)
                    
                    # 检查是否第一列为主键
                    has_pk = any('PRIMARY KEY' in ct for ct in column_types[:1]) if column_types else False
                    
                    # 创建表结构
                    db.create_table(table_name, headers, column_types, has_pk)
            
            # 获取合并单元格信息（仅Excel文件）
            merged_ranges = []
            if file_type == 'excel':
                merged_ranges = get_merged_cells_info(file_path, current_sheet)
                if merged_ranges:
                    logger.info(f"检测到 {len(merged_ranges)} 个合并单元格区域，将自动处理")
            
            # 计算块数
            total_chunks = (estimated_rows + chunk_size - 1) // chunk_size
            logger.info(f"将分成 {total_chunks} 个数据块并行处理")
            
            pbar = tqdm(total=estimated_rows, desc=f"处理 {current_sheet}")
            
            # 在file_to_database函数中的"准备并行任务"部分
            # 准备并行任务
            tasks = []
            for chunk_id in range(total_chunks):
                # 计算跳过的行数 - CSV和Excel处理稍有不同
                skiprows = chunk_id * chunk_size  # 跳过已处理行
                if file_type == 'excel' or (file_type == 'csv' and csv_props['has_header']):
                    skiprows += 1  # 额外跳过表头行
                
                tasks.append((chunk_id, file_path, current_sheet, skiprows, chunk_size, 
                            headers, merged_ranges, file_type, csv_props, transform_rules))
            
            processed_rows = 0
            with concurrent.futures.ProcessPoolExecutor(max_workers=max_workers) as executor:
                futures = {
                    executor.submit(process_chunk, t): t[0]
                    for t in tasks
                }
                data_chunks = []
                for future in concurrent.futures.as_completed(futures):
                    chunk_id, chunk_data = future.result()
                    if chunk_data:
                        # 按块ID有序插入
                        insert_pos = 0
                        for i, (existing_id, _) in enumerate(data_chunks):
                            if chunk_id < existing_id:
                                insert_pos = i
                                break
                            else:
                                insert_pos = i + 1  # 修复：将i + a = 1改为i + 1
                        data_chunks.insert(insert_pos, (chunk_id, chunk_data))
                        
                        processed_rows += len(chunk_data)
                        pbar.update(len(chunk_data))
                    
                    if len(data_chunks) % 5 == 0:
                        gc.collect()
            
            pbar.close()
            
            # 整理后写入数据库
            chunk_data_only = [chunk for _, chunk in data_chunks]
            logger.info("将处理好的数据写入数据库...")
            total_inserted = db.write_data(table_name, headers, chunk_data_only, excel_to_db_mapping)
            logger.info(f"成功插入 {total_inserted} 行数据到表 '{table_name}'")
            
            # 创建索引（只在覆盖模式或表不存在时）
            if mode == 'overwrite' or not table_exists:
                db.create_indices(table_name, headers, max_indices)
            
            sheet_time = time.time() - sheet_start_time
            results.append({
                'sheet_name': current_sheet,
                'table_name': table_name,
                'rows_processed': processed_rows,
                'time_seconds': sheet_time
            })
            
            total_rows_processed += processed_rows
            
            logger.info(f"工作表 '{current_sheet}' 处理完成! 耗时: {sheet_time:.2f} 秒")
            
            data_chunks = None
            chunk_data_only = None
            gc.collect()
        
        # 全部完成后优化数据库
        db.optimize()
        
        # 验证数据库
        db.verify()
        
        # 断开连接
        db.disconnect()
        
        end_time = time.time()
        total_time = end_time - start_time
        
        logger.info(f"处理完成! 总共处理 {total_rows_processed} 行数据")
        logger.info(f"总耗时: {total_time:.2f} 秒")
        
        if total_time > 0:
            logger.info(f"平均处理速度: {total_rows_processed / total_time:.2f} 行/秒")
        
        return {
            'success': True,
            'file_path': file_path,
            'file_type': file_type,
            'db_type': db_type,
            'total_sheets': len(sheets_to_process),
            'total_rows': total_rows_processed,
            'time_seconds': total_time,
            'sheet_results': results
        }
    
    except Exception as e:
        logger.error(f"处理过程中出错: {str(e)}", exc_info=True)
        return {
            'success': False,
            'error': str(e)
        }

# 用于MySQL密码输入的工具函数
def get_mysql_password(args):
    """获取MySQL密码，如果未在命令行提供则交互式提示输入"""
    if args.mysql_password:
        return args.mysql_password
    elif args.db_type == 'mysql':
        return getpass.getpass("请输入MySQL密码: ")
    return None

# 字段映射功能
def load_field_mapping(mapping_file):
    """从JSON或CSV文件加载字段映射配置"""
    logger = logging.getLogger(__name__)
    
    try:
        if not os.path.exists(mapping_file):
            logger.error(f"映射文件不存在: {mapping_file}")
            return None
            
        file_ext = os.path.splitext(mapping_file)[1].lower()
        
        if file_ext == '.json':
            # JSON格式: {sheet_name: {excel_column: db_column, ...}, ...}
            import json
            with open(mapping_file, 'r', encoding='utf-8') as f:
                mapping = json.load(f)
                
            # 验证格式
            if not isinstance(mapping, dict):
                logger.error("映射文件格式错误: 应为字典格式")
                return None
                
            logger.info(f"已加载JSON映射文件: {mapping_file}")
            return mapping
            
        elif file_ext == '.csv':
            # CSV格式: sheet_name,excel_column,db_column
            mapping = {}
            
            with open(mapping_file, 'r', encoding='utf-8') as f:
                import csv
                reader = csv.reader(f)
                header = next(reader)  # 跳过表头
                
                for row in reader:
                    if len(row) < 3:
                        continue
                        
                    sheet_name, excel_col, db_col = row[0], row[1], row[2]
                    
                    if sheet_name not in mapping:
                        mapping[sheet_name] = {}
                        
                    mapping[sheet_name][excel_col] = db_col
            
            logger.info(f"已加载CSV映射文件: {mapping_file}")
            return mapping
            
        else:
            logger.error(f"不支持的映射文件格式: {file_ext}")
            return None
            
    except Exception as e:
        logger.error(f"加载映射文件出错: {str(e)}")
        return None
        
def parse_inline_mapping(mapping_str):
    """解析内联映射字符串 'sheet:excol1=dbcol1,excol2=dbcol2;sheet2:...'"""
    logger = logging.getLogger(__name__)
    
    try:
        mapping = {}
        
        # 按工作表分割
        sheet_mappings = mapping_str.split(';')
        
        for sheet_mapping in sheet_mappings:
            if not sheet_mapping.strip():
                continue
                
            # 分离工作表名和映射
            parts = sheet_mapping.strip().split(':', 1)
            if len(parts) != 2:
                logger.warning(f"跳过无效的映射: {sheet_mapping}")
                continue
                
            sheet_name, column_mappings = parts
            sheet_name = sheet_name.strip()
            
            if sheet_name not in mapping:
                mapping[sheet_name] = {}
                
            # 处理列映射
            column_pairs = column_mappings.split(',')
            for pair in column_pairs:
                if not pair.strip():
                    continue
                    
                col_parts = pair.split('=', 1)
                if len(col_parts) != 2:
                    logger.warning(f"跳过无效的列映射: {pair}")
                    continue
                    
                excel_col, db_col = col_parts[0].strip(), col_parts[1].strip()
                mapping[sheet_name][excel_col] = db_col
        
        logger.info(f"已解析内联映射配置，包含 {len(mapping)} 个工作表")
        return mapping
        
    except Exception as e:
        logger.error(f"解析内联映射出错: {str(e)}")
        return None

def apply_column_transformation(excel_value, transform_rule):
    """应用字段转换规则"""
    if not transform_rule or excel_value is None:
        return excel_value
        
    # 转换规则示例: 'uppercase', 'lowercase', 'trim', 'prefix:ABC', 'suffix:XYZ', 'replace:old,new'
    rule_parts = transform_rule.split(':', 1)
    rule_name = rule_parts[0].lower()
    
    if rule_name == 'uppercase':
        return str(excel_value).upper()
    elif rule_name == 'lowercase':
        return str(excel_value).lower()
    elif rule_name == 'trim':
        return str(excel_value).strip()
    elif rule_name == 'prefix' and len(rule_parts) > 1:
        return f"{rule_parts[1]}{excel_value}"
    elif rule_name == 'suffix' and len(rule_parts) > 1:
        return f"{excel_value}{rule_parts[1]}"
    elif rule_name == 'replace' and len(rule_parts) > 1:
        old, new = rule_parts[1].split(',', 1)
        return str(excel_value).replace(old, new)
    elif rule_name == 'date_format' and len(rule_parts) > 1:
        from datetime import datetime
        try:
            # 假设excel_value是日期对象
            return excel_value.strftime(rule_parts[1])
        except:
            return excel_value
    else:
        return excel_value

# 解析列转换规则
def parse_column_transform(transform_str):
    """解析列转换规则字符串 'sheet:column:rule;...'"""
    logger = logging.getLogger(__name__)
    
    try:
        transforms = {}
        
        if not transform_str:
            return transforms
            
        # 按分号分割不同的转换规则
        rule_items = transform_str.split(';')
        
        for item in rule_items:
            if not item.strip():
                continue
                
            # 分割工作表、列和规则
            parts = item.strip().split(':', 2)
            if len(parts) != 3:
                logger.warning(f"跳过无效的转换规则: {item}")
                continue
                
            sheet_name, column_name, rule = [p.strip() for p in parts]
            
            if not sheet_name or not column_name or not rule:
                logger.warning(f"跳过空的转换规则: {item}")
                continue
                
            if sheet_name not in transforms:
                transforms[sheet_name] = {}
                
            transforms[sheet_name][column_name] = rule
        
        logger.info(f"已解析列转换规则，共 {len(transforms)} 个工作表")
        return transforms
        
    except Exception as e:
        logger.error(f"解析列转换规则出错: {str(e)}")
        return {}

# 命令行入口
def main():
    parser = argparse.ArgumentParser(description='通用Excel/CSV到SQLite/MySQL转换工具')
    
    parser.add_argument('file_path', nargs='?', help='输入文件路径 (Excel/CSV)')
    
    # 数据库类型和参数
    parser.add_argument('--db-type', choices=['sqlite', 'mysql'], default='sqlite',
                        help='目标数据库类型 (sqlite 或 mysql)')
    
    # SQLite参数
    parser.add_argument('--sqlite-path', help='SQLite数据库输出路径')
    
    # MySQL参数
    parser.add_argument('--mysql-host', default='localhost', help='MySQL主机地址')
    parser.add_argument('--mysql-port', type=int, default=3306, help='MySQL端口')
    parser.add_argument('--mysql-user', default='root', help='MySQL用户名')
    parser.add_argument('--mysql-password', help='MySQL密码 (不提供则交互式输入)')
    parser.add_argument('--mysql-database', help='MySQL数据库名')
    
    # CSV特定参数
    parser.add_argument('--csv-encoding', help='CSV文件编码 (默认自动检测)')
    parser.add_argument('--csv-separator', help='CSV分隔符 (默认自动检测)')
    parser.add_argument('--csv-quotechar', default='"', help='CSV引号字符 (默认: ")')
    parser.add_argument('--csv-no-header', action='store_true', help='CSV文件没有表头')
    
    # 通用参数
    parser.add_argument('-w', '--workers', type=int, default=None, 
                        help='工作进程数量 (默认为CPU核心数-1)')
    parser.add_argument('-c', '--chunk-size', type=int, default=20000, 
                        help='数据块大小 (默认: 20000)')
    parser.add_argument('-v', '--verbose', action='store_true', 
                        help='启用详细日志')
    parser.add_argument('-s', '--sheet', help='仅处理指定的工作表 (仅适用于Excel)')
    parser.add_argument('-l', '--list-sheets', action='store_true', 
                        help='仅列出Excel中的工作表')
    parser.add_argument('-i', '--max-indices', type=int, default=3,
                        help='每个表创建的最大索引数量')
    parser.add_argument('-q', '--quiet', action='store_true',
                        help='减少输出信息，仅显示关键进度')
    
    # 操作模式相关参数
    parser.add_argument('--mode', choices=['overwrite', 'append'], default='overwrite',
                        help='操作模式: overwrite=覆盖现有表, append=追加到现有表 (默认: overwrite)')
    parser.add_argument('--field-mode', choices=['create-all', 'match-only', 'mapping'], default='create-all',
                        help='字段处理模式: create-all=创建所有字段, match-only=仅处理目标已有字段, mapping=使用映射配置 (默认: create-all)')
    
    # 字段映射相关参数
    parser.add_argument('--mapping-file', help='字段映射配置文件路径(JSON或CSV格式)')
    parser.add_argument('--mapping', help='内联字段映射配置(格式:sheet:excol1=dbcol1,excol2=dbcol2;sheet2:...)')
    parser.add_argument('--column-transform', help='字段转换规则(格式:sheet:column:rule;...)，规则包括uppercase,lowercase,trim,prefix:X,suffix:X,replace:old,new,date_format:format')
    
    # 表名相关参数
    parser.add_argument('--target-table', help='指定目标数据库表名（覆盖默认的工作表名）')
    parser.add_argument('--table-mapping', help='工作表到数据库表的映射 (格式:excel_sheet=db_table,excel_sheet2=db_table2)')
    
    args = parser.parse_args()
    
    # 设置日志
    log_level = logging.ERROR if args.quiet else (logging.DEBUG if args.verbose else logging.INFO)
    
    # 必要依赖检查
    try:
        import chardet
    except ImportError:
        print("错误: 缺少必要的依赖项 'chardet'")
        print("请执行以下命令安装: pip install chardet")
        sys.exit(1)
    
    # 如果只想列出工作表
    if args.list_sheets:
        if not args.file_path:
            print("错误: 未指定输入文件路径")
            sys.exit(1)
            
        setup_logger(log_level)
        
        file_type = detect_file_type(args.file_path)
        if file_type != 'excel':
            print("错误: --list-sheets 选项仅适用于Excel文件")
            sys.exit(1)
            
        try:
            excel_info = get_excel_info(args.file_path)
            print("Excel文件包含以下工作表:")
            for i, name in enumerate(excel_info['sheet_names'], 1):
                print(f"{i}. {name}")
            sys.exit(0)
        except Exception as e:
            print(f"错误: {str(e)}")
            sys.exit(1)
    
    # 必要参数检查
    if not args.file_path:
        parser.print_help()
        sys.exit(1)
    
    if not os.path.exists(args.file_path):
        print(f"错误: 输入文件 '{args.file_path}' 不存在")
        sys.exit(1)
    
    # 根据数据库类型验证参数
    if args.db_type == 'sqlite':
        if not args.sqlite_path:
            print("错误: 使用SQLite时必须指定 --sqlite-path 参数")
            sys.exit(1)
        
        # 设置日志文件
        log_file = f"{os.path.splitext(args.sqlite_path)[0]}_conversion.log"
        setup_logger(log_level, log_file)
        
        # 构建SQLite参数
        db_params = {
            'db_path': args.sqlite_path
        }
    
    elif args.db_type == 'mysql':
        if not args.mysql_database:
            print("错误: 使用MySQL时必须指定 --mysql-database 参数")
            sys.exit(1)
        
        # 获取MySQL密码
        mysql_password = get_mysql_password(args)
        
        # 设置日志文件
        log_file = f"mysql_{args.mysql_database}_conversion.log"
        setup_logger(log_level, log_file)
        
        # 构建MySQL参数
        db_params = {
            'host': args.mysql_host,
            'port': args.mysql_port,
            'user': args.mysql_user,
            'password': mysql_password,
            'database': args.mysql_database
        }
    
    # 解析字段映射
    field_mapping = None
    if args.mapping_file:
        field_mapping = load_field_mapping(args.mapping_file)
    elif args.mapping:
        field_mapping = parse_inline_mapping(args.mapping)
    
    # 解析字段转换规则
    column_transforms = None
    if args.column_transform:
        column_transforms = parse_column_transform(args.column_transform)
    
    # 解析表映射参数
    table_mapping = None
    if args.table_mapping:
        table_mapping = parse_table_mapping(args.table_mapping)
    
    # 解析CSV特定参数
    csv_params = None
    file_type = detect_file_type(args.file_path)
    if file_type == 'csv':
        # 如果提供了CSV特定参数，构建参数字典
        if args.csv_encoding or args.csv_separator or args.csv_no_header:
            csv_params = {}
            
            # 如果指定了编码，使用指定的；否则自动检测
            if args.csv_encoding:
                csv_params['encoding'] = args.csv_encoding
            else:
                # 检测编码
                with open(args.file_path, 'rb') as f:
                    sample = f.read(4096)
                result = chardet.detect(sample)
                csv_params['encoding'] = result['encoding'] if result['confidence'] > 0.7 else 'utf-8'
            
            # 如果指定了分隔符，使用指定的；否则自动检测
            if args.csv_separator:
                csv_params['sep'] = args.csv_separator
            else:
                # 简单检测分隔符
                with open(args.file_path, 'r', encoding=csv_params['encoding'], errors='replace') as f:
                    sample = f.read(4096)
                    comma_count = sample.count(',')
                    semicolon_count = sample.count(';')
                    tab_count = sample.count('\t')
                    
                    if tab_count > comma_count and tab_count > semicolon_count:
                        csv_params['sep'] = '\t'
                    elif semicolon_count > comma_count:
                        csv_params['sep'] = ';'
                    else:
                        csv_params['sep'] = ','
            
            # 设置引号字符
            csv_params['quotechar'] = args.csv_quotechar
            
            # 设置是否有表头
            csv_params['has_header'] = not args.csv_no_header
    
    # 执行转换
    result = file_to_database(
        file_path=args.file_path,
        db_type=args.db_type,
        db_params=db_params,
        sheet_name=args.sheet,
        chunk_size=args.chunk_size,
        max_workers=args.workers,
        max_indices=args.max_indices,
        mode=args.mode,
        field_mode=args.field_mode,
        field_mapping=field_mapping,
        column_transforms=column_transforms,
        target_table=args.target_table,
        table_mapping=table_mapping,
        csv_params=csv_params
    )
    
    if result['success']:
        print("\n转换成功完成!")
        if args.db_type == 'sqlite':
            print(f"SQLite数据库已保存到: {args.sqlite_path}")
        else:
            print(f"数据已导入到MySQL数据库: {args.mysql_database}")
            
        print(f"总共处理了 {result['total_rows']} 行数据，耗时 {result['time_seconds']:.2f} 秒")
        print(f"文件类型: {result['file_type'].upper()}")
        
        if len(result['sheet_results']) > 1:
            print("\n各工作表处理结果:")
            for sheet_result in result['sheet_results']:
                print(f"- {sheet_result['sheet_name']}: {sheet_result['rows_processed']} 行，"
                      f"耗时 {sheet_result['time_seconds']:.2f} 秒")
    else:
        print(f"\n转换失败: {result['error']}")
        sys.exit(1)

if __name__ == "__main__":
    main()