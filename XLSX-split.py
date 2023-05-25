import csv
import os
import argparse
import configparser
import codecs
from openpyxl import load_workbook

# 解析命令行参数
parser = argparse.ArgumentParser(description='Process XLSX data and generate CSV files.')
parser.add_argument('-c', '--config', default='config.ini', help='path to the configuration file')
args = parser.parse_args()

# 读取配置文件
config = configparser.ConfigParser()
with codecs.open(args.config, 'r', encoding='utf-8') as file:
    config.read_file(file)

# 获取配置信息
xlsx_file = config.get('General', 'xlsx_file')
output_directory = config.get('General', 'output_directory')
raw_sheet_name = config.get('General', 'raw_sheet_name')
csv_encoding = config.get('General', 'csv_encoding', fallback='utf-8')
department_column_name = config.get('General', 'KEY', fallback='二级部门名称')

tag_departments = {}
for tag, departments in config.items('TagDepartments'):
    tag_departments[tag] = [dep.strip() for dep in departments.split(',')]

# 读取XLSX文件
workbook = load_workbook(xlsx_file)

# 检查是否存在名为 raw_sheet_name 的子表
if raw_sheet_name in workbook.sheetnames:
    sheet = workbook[raw_sheet_name]

    # 读取表格数据
    data = []
    columns_to_process = config.get('ColumnMappings', raw_sheet_name, fallback="姓名, 邮箱前缀, 一级部门名称, 二级部门名称").split(",")
    print("列名称:", columns_to_process)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for index, column in enumerate(columns_to_process):
            row_data[column] = row[index]

        data.append(row_data)

    # 添加标签字段
    for row in data:
        for tag, departments in tag_departments.items():
            if row[department_column_name] in departments:
                row['标签'] = tag
                break
        else:
            row['标签'] = '其他'


    # 根据标签生成多个CSV表
    tags = set(row['标签'] for row in data)
    for tag in tags:
        filename = f"{tag}.csv"
        filepath = os.path.join(output_directory, filename)
        tag_data = [row for row in data if row['标签'] == tag]

        with open(filepath, mode='w', newline='', encoding=csv_encoding) as file:
            writer = csv.DictWriter(file, fieldnames=tag_data[0].keys())
            writer.writeheader()
            writer.writerows(tag_data)


        print(f"已生成CSV文件: {filepath}")
else:
    print(f"名为 '{raw_sheet_name}' 的子表不存在。")
